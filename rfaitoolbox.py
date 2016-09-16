# Toolbox with functions used in the RFAI tools (PTA + Admin)
# last revision: Sep 2016, Michael Fruhnert + Katherine Rothe

# Toolbox is used for rfaiPTA.py AND rfaiAdmin.py

#import all necessary libraries
import setupDirAndStuff as dir14x 
import basicUtils as basics
from openpyxl import load_workbook
from openpyxl.cell.cell import get_column_letter
from os import system, remove, makedirs
from os.path import isfile, exists
import os
import unicodedata
from unidecode import unidecode

##identify global variables - used across the scripts
term = basics.getTerm(); # adapt to Fall, Spring and year (to have an archive)
# name of excel workbook with questions and answers
nameDB = dir14x.gsRFAIData + "{:s}_{:s}{:d}.xlsx".format('RFAI', term[0], term[3])
qWsName = "questions" #name of question sheet in workbook
catWsName = "categories" #name of categories sheet in workbook
qRowsXLS = [30, 34, 38, 42, 46] #rows of questions in xlsx template
roster = basics.loadRoster() # has a list of teams and students, use to find them fast

# function looks for first RFAI folders in Current
# if ADMIN = True, then it also looks in Upcoming        
def findRFAIFolder(admin = False):
    for item in os.listdir(dir14x.assnEntryDir):
        if ('RFAI' in item):
            return [os.path.join(dir14x.assnEntryDir, item), item]
            
    if admin:
        for item in os.listdir(dir14x.assnUpcoming):
            if ('RFAI' in item):
                return [os.path.join(dir14x.assnUpcoming, item), item]
            
    print("No RFAI found")
    return ['', '']
    

#finds row number of indicated question (by ID)
#returns row number if found, returns 0 if not
#qWs: question worksheet, findThis: qID
def getRowNum(qWs, findThis):
    qRow = 1
    myCell = qWs['A' + str(qRow)]
    while (myCell.value != None):
        if(myCell.value == findThis):
            return qRow    #returns row number if found
        qRow += 2 # assume fixed structure, one "empty line" between qIDs
        myCell = qWs['A' + str(qRow)]
    return 0    #indicates row not found

#gets all question categories
def getCategories():
    wb = load_workbook(nameDB) # load every time to get latest version
    catWs = wb.get_sheet_by_name(catWsName)   #gets categories sheet
    cats = []
    i = 1 #fixed structure, empty line: stop
    myCell = catWs['A' + str(i)]
    while (myCell.value != None):
        cats.append(str(myCell.value))
        i += 1
        myCell = catWs['A' + str(i)]
    return cats

# function translates column number into 'A2', 'AZ1' excel format
def cellID(col, row):
    return (get_column_letter(col) + str(row)) # col 1 = 'A'
    
# function returns list of entries for specified row and worksheet
def getRow(ws, row, offset = 1): # offset = 1: start at 'B'
    content = []
    i = offset + 1 #fixed structure, empty line: stop
    myCell = ws[cellID(i, row)]
    while (myCell.value != None):
        content.append(str(myCell.value))
        i += 1
        myCell = ws[cellID(i, row)]
    return content    
    
# prints questions, flavours and answer based on row number
def printQ(qWs, qRow):
    qID = str(qWs['A' + str(qRow)].value)
    # qID[1:] takes 'Q' of
    print('\033[1m' + "{:2d}: ".format(int(qID[1:])) + ' | '.join(getRow(qWs, qRow)))
    print('\033[0m' + "   A: " + str(qWs['B' + str(qRow + 1)].value) + '\n')
    
#function prints the questions with a specific category label
def printCat(catRow):
    wb = load_workbook(nameDB) # load every time to get latest version
    qWs = wb.get_sheet_by_name(qWsName) # gets Q&A sheet
    catWs = wb.get_sheet_by_name(catWsName)   #gets categories sheet
    
    for qID in getRow(catWs, catRow):
        qRow = getRowNum(qWs, qID)
        if (qRow != 0): # skip invalid or deleted questions
            printQ(qWs, qRow)   

#insert question as new, need to lock master
def addNewQ(stuQuestion):
    if (lockMaster() == 1): #try to lock master
        wb = load_workbook(nameDB) # load every time to get latest version
        qWs = wb.get_sheet_by_name(qWsName) # gets Q&A sheet
        # find last ID
        qRow = 1
        myCell = qWs['A' + str(qRow)]
        while (myCell.value != None):
            qRow += 2 # assume fixed structure, one "empty line" between qIDs
            myCell = qWs['A' + str(qRow)]
        
        if (qRow == 1): # if no question ever entered before
            qID = 'Q1'
        else:
            qID_last = str(qWs['A' + str(qRow - 2)].value)
            qID = 'Q' + str(int(qID_last[1:]) + 1)
        
        # add new ID and question
        qWs['A' + str(qRow)].value = qID
        qWs['B' + str(qRow)].value = stuQuestion
        qWs['B' + str(qRow + 1)].value = '[NEW]'
        wb.save(nameDB)
        unlockMaster()
        return qID
    else:
        return None
        
# browse through categories or list all questions
def searchCats(stuQuestion):
    system('clear')
    while (True):
        if (stuQuestion != ''): # do not show in admin mode
            print('Q_asked: ' + stuQuestion) # print for reference
        categories = getCategories()
        for ind, item in enumerate(categories):
            print("{:2d}: ".format(ind + 1) + item)
        doThis = input("\nEnter category (1, 2, ...) or\n" \
                        "0: return\n" \
                        "\n")
                        
        if (doThis.isdigit()):
            doThis = int(doThis)
        else:
            doThis = -1

        if (doThis == 0):
            return
        else:
            if ((doThis < 0) or (doThis > len(categories))):
                print("\nCategory invalid. Try again.\n")
            else:
                system('clear')
                printCat(doThis)
                return

# show all answers
def showAllAnswers():
    system('clear')
    wb = load_workbook(nameDB) # load every time to get latest version
    qWs = wb.get_sheet_by_name(qWsName) # gets Q&A sheet
    qRow = 1 
    qID = qWs['A' + str(qRow)].value
    while (qID != None):
        if(qID[0] == 'Q'): # skip deleted questions still in system
            printQ(qWs, qRow)
        qRow += 2 # assume fixed structure, one "empty line" between qIDs
        qID = qWs['A' + str(qRow)].value
    return

# function searches database by keywords
# 0 (1) - Do not (Do) include Answer/Keyword row                    
def searchByKeyword(includeAnswers):
    keys = input("Enter keyword: ").lower().split()
    qRow = 1
    wb = load_workbook(nameDB) # load every time to get latest version
    qWs = wb.get_sheet_by_name(qWsName) # gets Q&A sheet
    qID = str(qWs['A' + str(qRow)].value)
    while (qID != 'None'): #iterate through qIDs
        cellVal = ''
        if (qID[0] == 'Q'): #skip removed questions
            for i in range(includeAnswers + 1):
                cellVal += (' '.join(getRow(qWs, qRow + i))).lower()
        
        allKeysIn = True
        for key in keys:
            allKeysIn = (allKeysIn and (key in cellVal))
        if (allKeysIn):
            printQ(qWs, qRow)                
        
        qRow += 2 # assume fixed structure, one "empty line" between qIDs
        qID = str(qWs['A' + str(qRow)].value)    
                    
# sub-function to assign / edit questions - returns qID
# ws is passed as reference, not value - edits propagate
def editQuestion(rubricPath, rubricWs, rubricQ):
    if (rubricPath != ''): # if called from Admin tool
        stuQuestion = str(rubricWs['B' + str(qRowsXLS[rubricQ - 1])].value)
        insert1 = "\nEnter ANSWER ID (1, 2, ...) or\n"
        insert2 = "r: raise (add) as new question\n"
    else:
        if (rubricQ != ''):
            insert1 = "\nEnter question ID (1, 2, ...) to replace with or\n"
        else:
            insert1 = "\nEnter question ID (1, 2, ...) to be replaced or\n"
        stuQuestion = rubricQ
        insert2 = ''
        
    system('clear')
    qMatch = None
    while (qMatch == None):
        if (stuQuestion != ''):
            print('Q_asked: ' + stuQuestion)
        doThis = input(insert1 + \
                        "a: show all answers\n" \
                        "b: browse categories\n" \
                        "s: search by keyword\n" \
                        + insert2 + \
                        "0: return\n" \
                        "\n")
        if (doThis == 'a'):
            showAllAnswers()
        elif (doThis == 'b'):
            searchCats(stuQuestion)
        elif (doThis == 's'):
            searchByKeyword(1) # 1 - include Answers + Keywords
        elif ((doThis == 'r') and (rubricPath != '')):
            qMatch = addNewQ(stuQuestion)
        else:
            if (doThis.isdigit()):
                doThis = int(doThis)
            else:
                print('Invalid entry. Enter 0 to return')
                continue
                
            if (doThis == 0):
                qMatch = ''
            else:
                wb = load_workbook(nameDB) # load every time to get latest version
                qWs = wb.get_sheet_by_name(qWsName)   #gets questions sheet
                qRow = getRowNum(qWs, 'Q'  + str(doThis))
                if (qRow == 0):
                    print("\nID invalid. Try again.\n")
                else:
                    system('clear')
                    qMatch = ('Q'  + str(doThis)) # return qID
    
    if (qMatch != ''):
        if (rubricPath != ''): # if not called from admin
            rubricWb = load_workbook(rubricPath) # pull most recent version
            rubricWs = rubricWb.active
            rubricWs['A' + str(qRowsXLS[rubricQ - 1])].value = qMatch
            rubricWb.save(rubricPath)
        else:
            rubricWs = [qMatch, str(qWs['B' + str(qRow)].value), qRow]
    else:
        if (rubricPath == ''):
            rubricWs = ['', '', 0]
        
    return rubricWs
    
# looks for rubrics / retrieves file (same as GS update script)
def retrieveRubric(pathToRFAI, teamNum):
    sectionNum = basics.getSection(roster[2], teamNum)
    teamFolderPath = basics.getTeamFolderPath(sectionNum, teamNum)
    assn = basics.getFolderName(pathToRFAI)
    rubricPath = os.path.join(pathToRFAI, teamFolderPath,
                    '{}_Team{:02d}.xlsx'.format(assn, teamNum))
    if isfile(rubricPath):
        return rubricPath
    else:
        return '' # assignment is missing
        
# display answer based on qID            
def mapAns(qWs, qID):
    if (qID == 'None'): # not defined yet
        return '[SELECT ANSWER]'
    qRow = getRowNum(qWs, qID)
    if(qRow == 0):
        return '[NOT FOUND]'    #unsuccessful completion
    answer = str(qWs['B' + str(qRow + 1)].value)
    return answer.replace('\\n', '\n')    #successful completion
            
# display questions and answers for the selected rubric / worksheet            
def displayRubric(rubricWs, rfaiName):
    print(str(rubricWs['B18'].value) + ', ' + rfaiName + 
                ' - PTA done: ' + str(rubricWs['ZZ1'].value == True))
    wb = load_workbook(nameDB) # load every time to get latest version
    qWs = wb.get_sheet_by_name(qWsName) # gets Q&A sheet
    for i in range(5):
        print('\033[1m' + 'Q' + str(i + 1) + ': ' + str(rubricWs['B' + str(qRowsXLS[i])].value))
        print('\033[0m' + '  A: ' + mapAns(qWs, str(rubricWs['A' + str(qRowsXLS[i])].value)))
        print(' ')
        
# lock master spreadsheet for edits
def lockMaster():
    filePath = os.path.join(dir14x.assnEntryDir, 'lockRFAIMaster.txt')
    if isfile(filePath):
        print("Master cannot be locked. Try again later")
        return 0
    else:
        notifier = open(filePath, 'w').close()
        return 1
    
# remove lock (safe changes before)   
def unlockMaster():
    remove(os.path.join(dir14x.assnEntryDir, 'lockRFAIMaster.txt'))
    
    
########################################################################
# Functions for rfai Admin only implemented below
########################################################################

# function overwrites a row in the master sheet
def setRow(wsName, entries, row, offset = 1):
    if (lockMaster() == 1): #try to lock master
        wb = load_workbook(nameDB) # load every time to get latest version
        ws = wb.get_sheet_by_name(wsName) # get sheet

        for ind, item in enumerate(entries): #ind starts at 0, +1 for col 'A'
            ws[cellID(ind + offset + 1, row)].value = item
            
        wb.save(nameDB)
        unlockMaster()
        print("Update successful\n")
        return True
    return False # false if failed

# extended printQ that lists keywords and categories
def printQExt(qRow):
    wb = load_workbook(nameDB) # load every time to get latest version
    qWs = wb.get_sheet_by_name(qWsName) # gets Q&A sheet
    catWs = wb.get_sheet_by_name(catWsName) # gets category sheet
    qID = str(qWs['A' + str(qRow)].value)
    print('\033[1m' + "{:3d} : ".format(int(qID[1:])) + ' | '.join(getRow(qWs, qRow)))
    print('\033[0m' + "Ans : " + str(qWs['B' + str(qRow + 1)].value))
    print('\033[1m' + "Keys: " + ' | '.join(getRow(qWs, qRow + 1, 2)))
    catsListed = []
    for ind, item in enumerate(getCategories()):
        qIDsInCat = ' '.join(getRow(catWs, ind + 1))
        if (qID in qIDsInCat):
            catsListed.append(item)
    print('\033[0m' + "Cats: "+ ' | '.join(catsListed) + "\n")

# procedure allows modification of questions, flavours and keywords
def modifyQEntry(qRow):
    system('clear')
    while (1):
        wb = load_workbook(nameDB) # load every time to get latest version
        qWs = wb.get_sheet_by_name(qWsName) # gets Q&A sheet
        selList = getRow(qWs, qRow)
        for ind, item in enumerate(selList):
            print("{:2d}".format(ind + 1) + ' : ' + item)

        doThis = input("\nEnter (1, 2, ...), 'a' to add, or 0 to exit:\n")
        
        system('clear')
        if (doThis == 'a'):
            selList.append(' ')
            doThis = str(len(selList))

        if (doThis.isdigit()):
            doThis = int(doThis)
        else:
            doThis = -1
            
        if (doThis == 0):
            return
        else:
            if ((doThis < 0) or (doThis > len(selList))):
                print('Invalid entry. Enter 0 to return')
            else:
                if (selList[doThis - 1] != ' '):
                    print("Replacing: " + selList[doThis - 1])
                newText = input("New formulation (leave blank to delete, use '\\n' for newline): \n").strip()
                if ((newText == '') and (doThis == 1)):
                    print('Cannot remove main question / answer')
                else:
                    if (newText == ''):
                        selList.pop(doThis - 1)
                        selList.append('')
                    else:
                        selList[doThis - 1] = newText
                    
                    setRow(qWsName, selList, qRow)
                    return
      
def changeCategory(catRow, qID, delete = True):
    wb = load_workbook(nameDB) # load every time to get latest version
    catWs = wb.get_sheet_by_name(catWsName) # gets category sheet
    selList = getRow(catWs, catRow)
    if delete:
        if (qID in selList):
            selList.remove(qID)
            selList.append('')
    else:
        selList.append(qID)
    return setRow(catWsName, selList, catRow)

def modifyCategories(qID):
    system('clear')
    while (1):
        wb = load_workbook(nameDB) # load every time to get latest version
        catWs = wb.get_sheet_by_name(catWsName) # gets category sheet
        
        categories = getCategories()
        catRowsListed = []
        catRowsNotListed = []
        for ind in range(len(categories)):
            qIDsInCat = ' '.join(getRow(catWs, ind + 1))
            if (qID in qIDsInCat):
                catRowsListed.append(ind + 1)
            else:
                catRowsNotListed.append(ind + 1)
        
        numActCats = len(catRowsListed) # number of active categories
        for ind in range(numActCats):
            print("{:2d} - rem: ".format(ind + 1) + categories[catRowsListed[ind] - 1])
        for ind in range(len(catRowsNotListed)):
            print("{:2d} - add: ".format(ind + numActCats + 1) + categories[catRowsNotListed[ind] - 1])
        
        doThis = input("\nPick category (1, 2, ...) to remove / add, or 0 to exit:\n")
        
        system('clear')
        if (doThis.isdigit()):
            doThis = int(doThis)
        else:
            doThis = -1
                
        if (doThis == 0):
            return
        else:
            if ((doThis < 0) or (doThis > len(catRowsListed) + len(catRowsNotListed))):
                print('Invalid entry. Enter 0 to return')
            else:
                if (doThis > len(catRowsListed)):
                    doThis -= len(catRowsListed)
                    print("Adding category: " + categories[catRowsNotListed[doThis - 1] - 1])
                    changeCategory(catRowsNotListed[doThis - 1], qID, False) # False - Do not delete
                    return
                else:
                    print("Removing category: " + categories[catRowsListed[doThis - 1] - 1])
                    changeCategory(catRowsListed[doThis - 1], qID)
                    return
      
# routine to edit / modify questions
def editQuestionMaster(qRow, qID):
    doMore = True
    system('clear')
    while(doMore):
        printQExt(qRow)

        doThis = input("Choose from options below:\n"\
                    "a: change answer (keywords)\n" \
                    "m: modify categories\n" \
                    "q: change question (flavors)\n" \
                    "0: return\n" \
                    "\n")
        
        system('clear')
        if (doThis == 'a'):
            modifyQEntry(qRow + 1)
        elif (doThis == 'm'):
            modifyCategories(qID)
        elif (doThis == 'q'):
            modifyQEntry(qRow)
        elif (doThis == '0'):
            doMore = False
        else:
            print("Invalid input. Enter 0 to exit.\n")
            
# sub-function to select / search for questions in master sheet
def selectQuestionMaster():
    system('clear')
    qMatch = None
    while (qMatch == None):
        doThis = input("\nEnter Question ID (1, 2, ...) or\n" \
                        "a: show all answer\n"
                        "b: browse categories\n"
                        "s: search by keyword\n"
                        "0: return\n" \
                        "\n")
        if (doThis == 'a'):
            showAllAnswers()
        elif (doThis == 'b'):
            searchCats('')
        elif (doThis == 's'):
            searchByKeyword(1) # 1 - include Answers + Keywords
        else:
            if (doThis.isdigit()):
                doThis = int(doThis)
            else:
                print('Invalid entry. Enter 0 to return')
                continue
                
            if (doThis == 0):
                qMatch = ''
            else:
                wb = load_workbook(nameDB) # load every time to get latest version
                qWs = wb.get_sheet_by_name(qWsName)   #gets questions sheet
                qRow = getRowNum(qWs, 'Q'  + str(doThis))
                if (qRow == 0):
                    print("\nID invalid. Try again.\n")
                else:
                    editQuestionMaster(qRow, 'Q'  + str(doThis)) # process with qID
                    qMatch = ''

def workWithCategory(catRow):
    system('clear')
    while (1):
        wb = load_workbook(nameDB) # load every time to get latest version
        catWs = wb.get_sheet_by_name(catWsName) # gets category sheet
        catEntries = getRow(catWs, catRow, 0) # 0 - retrieve 'header'
        print("category: " + catEntries[0])
        for ind in range(1, len(catEntries)):
            print("{:3d} : ".format(ind) + catEntries[ind])
        
        doThis = input("\nChoose from options below:\n"\
                           "c: change name\n" \
                           "r: re-order\n" \
                           "0: return\n" \
                           "\n")
        
        if (doThis == '0'):
            return
        elif (doThis == 'c'):
            newName = input("Enter new name: \n").strip()
            if (newName == ''):
                print("Cannot delete category.")
            else:
                catEntries[0] = newName
                setRow(catWsName, catEntries, catRow, 0)
            system('clear')
        elif (doThis == 'r'):
            qMove = int(input("Enter item (1, 2, ...) to move\n"))
            qBefore = int(input("Enter item (1, 2, ...) to place before\n"))
            system('clear')
            if ((qMove < 1) or (qBefore < 1) or (qMove >= len(catEntries)) or (qBefore >= len(catEntries)) or (qMove == qBefore)):
                print("Invalid moving parameters.\n")
            else:
                temp = catEntries[qMove]
                catEntries.pop(qMove)
                if (qMove < qBefore):
                    qBefore -= 1
                catEntries.insert(qBefore, temp)
                setRow(catWsName, catEntries, catRow, 0)
        else:
            system('clear')
            print('Invalid entry. Enter 0 to return')
                
# master function to rearrange categories                    
def rearrangeCategories():
    system('clear')
    while (1):
        wb = load_workbook(nameDB) # load every time to get latest version
        catWs = wb.get_sheet_by_name(catWsName) # gets category sheet
        
        categories = getCategories()
        for ind, item in enumerate(categories):
            print("{:2d} : ".format(ind + 1) + item)
        
        doThis = input("\nPick category (1, 2, ...), 'a' to add, or 0 to return:\n")
        
        system('clear')
        if (doThis == 'a'):
            setRow(catWsName, ['New Category'], len(categories) + 1, 0)
            continue
            
        if (doThis.isdigit()):
            doThis = int(doThis)
        else:
            doThis = -1
                
        if (doThis == 0):
            return
        else:
            if ((doThis < 0) or (doThis > len(categories) )):
                print('Invalid entry. Enter 0 to return')
            else:
                workWithCategory(doThis)
                system('clear')

# procedure to replace a question    
def replaceQuestion(pathToRFAI):
    totTeams = len(roster[1])
    qRemove = editQuestion('', None, '')
    qRemID = qRemove[0]
    if (qRemID != ''):
        qReplace = editQuestion('', None, qRemove[1])
        if (qReplace[0] != ''):
            print("Replacing " + qRemID + " by " + qReplace[0] + ".")
            if setRow(qWsName, [qReplace[0]], qRemove[2] + 1): # change answer
                categories = getCategories() # remove from all categories
                for i in range(len(categories)):
                    if not(changeCategory(i + 1, qRemID)):
                        return
                # order matters, once the next if statement executes properly, this cannot be undone
                if setRow(qWsName, ['E' + qRemID[1:], 'marked for deletion'], qRemove[2], 0):
                    for i in range(totTeams):
                        rubricPath = retrieveRubric(pathToRFAI, i + 1)
                        if (rubricPath != ''):
                            rubricWb = load_workbook(rubricPath) # pull most recent version
                            rubricWs = rubricWb.active
                            update = False
                            for qRow in qRowsXLS:
                                if (str(rubricWs['A' + str(qRow)].value) == qRemID):
                                    rubricWs['A' + str(qRow)].value = qReplace[0]
                                    update = True
                            if update:
                                rubricWb.save(rubricPath)
                    system('clear')

                    
########################################################################
# Statistics and file collection
########################################################################

# same as setRow, but assumes open master sheet
def setRowOpen(ws, entries, row, offset = 1):
        for ind, item in enumerate(entries): #ind starts at 0, +1 for col 'A'
            ws[cellID(ind + offset + 1, row)].value = item

# procedure to generate the statistics table
def collectStats():
    totTeams = len(roster[1])
    if (lockMaster() == 1):
        wb = load_workbook(nameDB) # load every time to get latest version
        qWs = wb.get_sheet_by_name(qWsName) # get Q&A sheet
        collectWs = wb.get_sheet_by_name('collection') # get collection sheet
        leaderWs = wb.get_sheet_by_name('stats') # get leaderboard sheet        
        
        qList = [] # generate list of questions
        qRow = 1
        qID = str(qWs['A' + str(qRow)].value)
        while (qID != 'None'): #iterate through qIDs
            if (qID[0] == 'Q'): #skip removed questions
                #qID, question, leader, RFAI 1, 2, ... 5
                newItem = [qID, str(qWs['B' + str(qRow)].value), 0, 0, 0, 0, 0, 0]
                qList.append(newItem)
            qRow += 2 # assume fixed structure, one "empty line" between qIDs
            qID = str(qWs['A' + str(qRow)].value) 
        
        # collect data
        for item in qList:
            for i in range(totTeams):
                total = ''
                for rfai in range(5):
                    cur = ' '.join(getRow(collectWs, i + 3, rfai * 6 + 1))
                    if (item[0] in cur):
                        item[3 + rfai] += 1
                    total += ' ' + cur
                if (item[0] in total):
                    item[2] += 1
        
        # generate leaderboard
        for rfai in range(5 + 1): #+1 for total
            qList.sort(key = lambda x: x[rfai + 2], reverse = True)
            while len(qList) < 10:
                qList.append([0] * 8) # add dummies
            for i in range(10): # get leading 10
                item = qList[i]
                if (item[2 + rfai] == 0): # do not list, but clear fields
                    item = [''] * 8                    
                leaderWs['A' + str(13 * rfai + 4 + i)].value = item[rfai + 2] # frequency
                leaderWs['B' + str(13 * rfai + 4 + i)].value = item[0] # ID
                leaderWs['C' + str(13 * rfai + 4 + i)].value = item[1] # question       
        
        wb.save(nameDB)
        unlockMaster()
    else:
        print("Could not collect statistics.")

# fills questions per team. Assumes empty collection table
def collectRFAI(pathToRFAI):
    totTeams = len(roster[1])
    assn = basics.getFolderName(pathToRFAI)
    rfaiNum = int(assn[-1]) # assume no more than 9 RFAIs per semester
    # check rubrics
    wb = load_workbook(nameDB) # load every time to get latest version
    qWs = wb.get_sheet_by_name(qWsName) # get Q&A sheet
    rubricList = []
    complete = True
    for i in range(totTeams):
        rubricPath = retrieveRubric(pathToRFAI, i + 1)
        if (rubricPath != ''):
            response = ('RFAI {0:d} Response\n\n' \
                        'Dear Team {1:02d},\n\n' \
                        'Please find the response to your request for additional information below.\n\n' \
                        'Kind regards,\n' \
                        'The Project Oversight Team').format(rfaiNum, i + 1)

            rubricWb = load_workbook(rubricPath) # pull most recent version
            rubricWs = rubricWb.active
            # try and catch a couple problems
            if not(rubricWs['ZZ1'].value == True):
                print('Team {:02d} not finished.'.format(i + 1))
                complete = False
            entries = []
            for qRow in qRowsXLS:
                qID = str(rubricWs['A' + str(qRow)].value)
                qAns = mapAns(qWs, qID)
                if ((qAns == '[NOT FOUND]') or (qAns == '[NEW]')):
                    print('Team {0:02d}, question {1:s} not found/answered.'.format(i + 1, qID))
                    complete = False
                if (qID != 'None'):
                    entries.append(qID)
                    response += '\n\nQ: ' + str(rubricWs['B' + str(qRow)].value) + \
                            '\nA: ' + '\n   '.join(qAns.split('\n'))
                            
            response = response.replace('\n', '\r\n') # change to Windows newline character
            rubricList.append([i + 1, response, entries])
    
    if complete:
        directory = os.path.join(pathToRFAI, 'FullRubrics')
        if not exists(directory):
            makedirs(directory)
            
        if (lockMaster() == 1):
            wb = load_workbook(nameDB) # load every time to get latest version
            collectWs = wb.get_sheet_by_name('collection') # get collection sheet
            
            for i in range(totTeams):
                collectWs['A' + str(i + 3)].value = 'Team {:02d}'.format(i + 1)
            
            for item in rubricList:
                teamNum = item[0]
                fileName = '/RFAI{0:d}_Response_Team{1:02d}.txt'.format(rfaiNum, teamNum)
                setRowOpen(collectWs, [''] * 6, teamNum + 2, rfaiNum * 6 - 5) #clean row
                setRowOpen(collectWs, item[2], teamNum + 2, rfaiNum * 6 - 5)
                with open(directory + fileName, 'w') as txtRubric:
                    txtRubric.write(item[1])
                    # automatic close
                
            wb.save(nameDB)
            unlockMaster()
            system('clear')
            print('Collection succesful. Generating statistics.')
            collectStats() # update statistics
            print('Statistics generated.\n')
    else:
        print('Process was terminated due to errors indicated above.\n')

########################################################################
# file cleaning - ASCII conformity
########################################################################        

# checks a single cell for ASCII conformity
def checkEntry(cellVal):
    valNew = unicodedata.normalize('NFKD', cellVal).encode('ascii', 'ignore').strip()
    if (valNew != cellVal):
        return [True, str(unidecode(cellVal).strip())]
    else:
        return [False, str(cellVal)]
        
# procedure checks for ASCII conform characters in single rubric       
def checkRubric(rubricPath):
    rubricWb = load_workbook(rubricPath) # pull most recent version
    rubricWs = rubricWb.active
    update = False
    for qRow in qRowsXLS:
        question = rubricWs['B' + str(qRow)].value
        if (question != None):
            asciiConform = checkEntry(question)
            if (asciiConform[0]):
                update = True
                rubricWs['B' + str(qRow)].value = asciiConform[1]

    if update:    
        rubricWb.save(rubricPath)
        print('Rubric updated to be ASCII conform.\n')

# checks row for ASCII conformity       
def checkRowOpen(ws, row, offset = 1): # offset = 1: start at 'B'
    content = []
    i = offset + 1 #fixed structure, empty line: stop
    myCell = ws[cellID(i, row)]
    update = False
    while (myCell.value != None):
        asciiConform = checkEntry(myCell.value)
        if (asciiConform[0]):
            print("Row: " + str(row) + " updated in col: " + str(i))
            update = True
        content.append(asciiConform[1])
        i += 1
        myCell = ws[cellID(i, row)]
        
    if update:
        setRowOpen(ws, content, row)
        
# procedure replaces all non-ascii conform characters        
def normalizeMaster():
    if (lockMaster() == 1):
        wb = load_workbook(nameDB) # load every time to get latest version
        qWs = wb.get_sheet_by_name(qWsName) # get Q&A sheet
        catWs = wb.get_sheet_by_name(catWsName) # get category sheet
        
        print("Checking questions sheet")
        qRow = 1
        qID = str(qWs['A' + str(qRow)].value)
        while (qID != 'None'): #iterate through qIDs
            checkRowOpen(qWs, qRow, 0) # questions
            checkRowOpen(qWs, qRow + 1) # answers / keywords
            qRow += 2 # assume fixed structure, one "empty line" between qIDs
            qID = str(qWs['A' + str(qRow)].value) 
            
        print("Checking categories sheet")
        catRow = 1
        myCell = catWs['A' + str(catRow)]
        while (myCell.value != None): #iterate through categories
            checkRowOpen(catWs, catRow, 0)
            catRow += 1 # assume fixed structure
            myCell = catWs['A' + str(catRow)]
        
        wb.save(nameDB)
        unlockMaster()
        print('Master spreadsheet normalized.\n')
            