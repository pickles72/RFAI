#! /usr/bin/env python3

# Script to stream-line RFAI processing
# last revision: Sep 2016, Michael Fruhnert + Katherine Rothe

## INSTRUCTIONS:
## 1)  Update the RFAI Master xlsx in GS / 10 RFAI Data
## 2a) The file with the RFAI questions and answers (defined by variable 'name'),
##      is formatted with 2 sheets with names: questions and categories
## 2b) The first column of the question sheet contains the question ID (QID)
##      in the form (Q##). QID are NOT to be altered
##      The second column contains the associated text (the question)
##      All following columns contain flavours of the same question
## 2c) The first column of the categories sheet contains the categories (text)
##     The second and following columns contain the associated QID
##     QIDs can be assigned to multiple categories
## 3)  To run, 'python3 /share/engr14x/rfai.py' in the command prompt
##      Follow the subsequent prompts to operate the program.
## 4)  To change global variables, like the database name, go to rfaitoolbox.py


#import all necessary libraries
from openpyxl import load_workbook
from os import system
from rfaitoolbox import retrieveRubric, checkRubric, displayRubric, editQuestion
from rfaitoolbox import findRFAIFolder
        
system("clear") # clears previous inputs
rfaiFolder = findRFAIFolder()
print(rfaiFolder[1] + ' entry launched.')

while 1:
    teamNum = int(input("Enter team number XX (0 to exit): "))
    if teamNum == 0:
        break
        
    rubric = retrieveRubric(rfaiFolder[0], teamNum)
    if (rubric == ''):
        print('\nTeam not found. Check if team actually submit.\n')
        continue
    else:
        system('clear')
        checkRubric(rubric)
        wb = load_workbook(rubric)
        ws = wb.active
        
    moreqs = True
    while(moreqs):
        displayRubric(ws, rfaiFolder[1])
        doThis = input("Enter question to edit (1, 2, ...) or\n" \
                    "0: return\n" \
                    "f: change finish mark\n" \
                    "\n")
        
        system("clear")
        if (doThis == 'f'):
            wb = load_workbook(rubric) # pull most recent version
            ws = wb.active
            ws['ZZ1'].value = not(ws['ZZ1'].value == True)
            wb.save(rubric)
        else:
            if (doThis.isdigit()):
                doThis = int(doThis)
            else:
                doThis = -1
                
            if ((doThis > 0) and (doThis < 6)):
                ws = editQuestion(rubric, ws, doThis)
            elif (doThis == 0):
                moreqs = False
            else:
                print("Invalid input. Enter c to cancel.\n")

print("RFAI-Tool shut-down...")